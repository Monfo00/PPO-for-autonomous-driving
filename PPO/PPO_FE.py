import numpy as np
import random
import os
os.environ['TF_ENABLE_ONEDNN_OPTS']='0'
import time
import tensorflow as tf
import cv2
import pandas as pd
from openpyxl import load_workbook
import random
from AirSimEnv import AirsimEnv
from tensorflow.keras.optimizers import Adam
import tensorflow_probability as tfp
from tensorflow.keras.layers import Conv2D, Flatten, Dense


INPUT_SHAPE = (66,200,3) 

weights_path_start = '.../main_drqn_model'
weights_path_actor = '.../actor.weights.h5'
weights_path_critic = '.../critic.weights.h5'
loss_file_path = '.../loss_history.xlsx'
reward_file_path = '.../reward_history.xlsx'
lr_actor=1e-4
lr_critic=1e-4
HORIZONT=320
BATCHSIZE=32
gamma=0.99
ppo_epochs=7
ppo_clip=0.2
tot_frame=1000000
GAMMA = 0.99
LAM = 0.95


STARTING_POINTS = [(88,-1,0.2,1,0,0,0),
                        (127.5,45,0.2,0.7,0,0,0.7),
                        (30,127.3,0.2,1,0,0,0),
                        (-59.5, 126,0.2,0,0,0,1),
                        (-127.2,28,0.2,0.7,0,0,0.7),
                        (-129, -48,0.2, 0.7, 0, 0, -0.7),
                        (-90, -128.5, 0.2, 0, 0,0,1),
                        (0,-86, 0.2, 0.7, 0,0, -0.7),
                        (62, -128.3, 0.2, 1, 0, 0, 0),
                        (127, -73, 0.2, 0.7, 0, 0, -0.7)]



class AirSimWrapper:

    def __init__(self, input_shape, ip, port):
        self.env = AirsimEnv(ip, port)
        self.input_shape = input_shape
        self.state = np.empty(input_shape)

    def frameProcessor(self, frame):
        frame = frame[40:136, 0:255, 0:3]
        frame = cv2.resize(frame, (self.input_shape[1], self.input_shape[0]), interpolation=cv2.INTER_NEAREST)
        return frame

    def reset(self, starting_point):

        observation = self.env.reset(starting_point)
        time.sleep(0.2)
        self.env.step(0,0)
        speed = self.env.client.getCarState().speed
        while speed < 3.0:
            speed = self.env.client.getCarState().speed

        frame = self.frameProcessor(observation)
        self.state = frame

        return frame

    def reset2(self, starting_point):

        observation = self.env.reset(starting_point)
        time.sleep(0.2)

        frame = self.frameProcessor(observation)
        self.state = frame

        return frame

    def step(self, action, j):

        new_frame, reward, done, info = self.env.step(action,j)
        processed_frame = self.frameProcessor(new_frame)

        self.state = processed_frame

        return processed_frame, reward, done

class Actor(tf.keras.Model):
    def __init__(self, action_size):
        super().__init__()

        self.action_size = action_size

        self.conv1 = Conv2D(32, (8, 8), activation='relu', strides=4, padding='valid', use_bias=False, trainable=False)
        

        self.conv2 = Conv2D(64, (4, 4), activation='relu', strides=2, padding='valid', use_bias=False, trainable=False)

        self.conv3 = Conv2D(64, (3, 3), activation='relu', strides=1, padding='valid', use_bias=False, trainable=False)

        self.flatten = Flatten()

        self.ad1 = Dense(512, activation='relu', kernel_initializer=tf.keras.initializers.Orthogonal(gain=np.sqrt(2)), bias_initializer=tf.keras.initializers.Constant(value=0))


        self.ad2_mean = Dense(1, activation='tanh', kernel_initializer=tf.keras.initializers.Orthogonal(gain=0.5),
                        bias_initializer=tf.keras.initializers.Constant(value=0))

        self.ad2_std = Dense(1, activation='softplus', kernel_initializer=tf.keras.initializers.Orthogonal(gain=0.5),
                        bias_initializer=tf.keras.initializers.Constant(value=0.5))


    def call(self, inputs):

        x = self.conv1(inputs)
        x = self.conv2(x)
        x = self.conv3(x)
        x = self.flatten(x)
        x = self.ad1(x)

        # Actor
        mean = self.ad2_mean(x)
        std = self.ad2_std(x)


        return mean, std
    

    def load_weights_model(self, filepath, path_valid):
        """Carica i pesi pre-addestrati solo per i layer convoluzionali."""
        dummy_input = tf.random.normal((1, 66, 200, 3))
        self(dummy_input)  # Costruisce la rete

        if os.path.exists(path_valid):
            self.load_weights(path_valid)
            print("Continua Training")
        else:
            if os.path.exists(filepath + ".index"):  # Controlla che il file di checkpoint esista
                try:
                    self.load_weights(filepath)
                    # Congela i layer convoluzionali
                    for layer in self.layers:
                        if isinstance(layer, tf.keras.layers.Conv2D) or isinstance(layer, tf.keras.layers.Flatten):
                            layer.trainable = False
                    print("Pesi CNN caricati con successo e congelati.")
                except Exception as e:
                    print(f"Errore nel caricamento dei pesi: {e}")
            else:
                print("File non trovato, inizializzo i pesi casualmente.")

        for layer in self.layers:
            print(f"Layer: {layer.name}, Trainable: {layer.trainable}")

    def save_weights_model(self, filepath):
        """Salva tutti i pesi del modello."""
        self.save_weights(filepath)

class Critic(tf.keras.Model):
    def __init__(self, action_size):
        super().__init__()

        self.action_size = action_size

        self.conv1 = Conv2D(32, (8, 8), activation='relu', strides=4, padding='valid', use_bias=False, trainable=False)

        self.conv2 = Conv2D(64, (4, 4), activation='relu', strides=2, padding='valid', use_bias=False, trainable=False)

        self.conv3 = Conv2D(64, (3, 3), activation='relu', strides=1, padding='valid', use_bias=False, trainable=False)



        self.flatten = Flatten()

        self.d1 = Dense(512, activation='relu', kernel_initializer=tf.keras.initializers.Orthogonal(gain=np.sqrt(2)), bias_initializer=tf.keras.initializers.Constant(value=0))


        self.val = Dense(1, kernel_initializer=tf.keras.initializers.Orthogonal(gain=0.5),
                            bias_initializer=tf.keras.initializers.Constant(value=0))


    def call(self, inputs):
        x = self.conv1(inputs)
        x = self.conv2(x)
        x = self.conv3(x)
        x = self.flatten(x)
        x = self.d1(x)


        # Critico
        val = self.val(x)
    
        return val

    def load_weights_model(self, filepath, path_valid):
        """Carica i pesi pre-addestrati solo per i layer convoluzionali."""
        dummy_input = tf.random.normal((1, 66, 200, 3))
        self(dummy_input)  # Costruisce la rete

        if os.path.exists(path_valid):
            self.load_weights(path_valid)
            print("Continua Training")
        else:
            if os.path.exists(filepath + ".index"):  # Controlla che il file di checkpoint esista
                try:
                    self.load_weights(filepath)
                    # Congela i layer convoluzionali
                    for layer in self.layers:
                        if isinstance(layer, tf.keras.layers.Conv2D) or isinstance(layer, tf.keras.layers.Flatten):
                            layer.trainable = False
                    print("Pesi CNN caricati con successo e congelati.")
                except Exception as e:
                    print(f"Errore nel caricamento dei pesi: {e}")
            else:
                print("File non trovato, inizializzo i pesi casualmente.")

    def save_weights_model(self, filepath):
        """Salva tutti i pesi del modello."""
        self.save_weights(filepath)



class Memory:
    def __init__(self):
        self.states = []
        self.actions = []
        self.rewards = []
        self.dones = []
        self.logprobs = []
        self.next_obs = []

   
    def forget(self):
        'Clear memory'
        self.states.clear()
        self.actions.clear()
        self.rewards.clear()
        self.dones.clear()
        self.logprobs.clear()
        self.next_obs.clear()





class PPO_Agent():
    def __init__(self, batch_size=BATCHSIZE, seed=321):

        np.random.seed(seed)  
        tf.random.set_seed(seed)  

        self.policy = Actor(action_size=1)
        self.policy.load_weights_model(weights_path_start, weights_path_actor)
        self.optimizer = Adam(learning_rate=lr_actor)
        self.policy.save_weights_model(weights_path_actor)

        self.policy_critic = Critic(action_size=1)
        self.policy_critic.load_weights_model(weights_path_start, weights_path_critic)
        self.optimizer2 = Adam(learning_rate=lr_critic)
        self.policy_critic.save_weights_model(weights_path_critic)

        self.policy_old = Actor(action_size=1)
        self.policy_old.load_weights_model(weights_path_start, weights_path_actor)

        self.MseLoss = tf.keras.losses.MeanSquaredError()
        self.batch_size = batch_size





    def select_action(self, state):

        mean, std = self.policy_old(state)
        #std = tf.clip_by_value(std, 0.15, 1.50)
        std = tf.clip_by_value(std, 0.15, 0.75)
        #dist = tfp.distributions.TruncatedNormal(loc=mean, scale=std, low=-1.00, high=1.00)
        dist = tfp.distributions.Normal(loc=mean, scale=std)
        actions = dist.sample()
        action_logprob = dist.log_prob(actions)
        #print(f"media stimata {mean} e azione {actions} e std {std}")
        return actions.numpy(), action_logprob.numpy()





    def update(self, memory, step):
        actor=[]
        critic=[]
        # Calcola i vantaggi
        T = len(memory.rewards)
        rewards = np.array(memory.rewards) 
        dones = np.array(memory.dones)

        values = []
        next_value = []
        for state, next_state in zip(memory.states, memory.next_obs):
            next_val = self.policy_critic(next_state)
            next_val = next_val.numpy()
            val = self.policy_critic(state)
            val = val.numpy()

            next_value.append(next_val)
            values.append(val)

        advantages = np.zeros(T)
        lastgaelam = 0
        for t in reversed(range(T)):
            if t == T - 1:
                nextnonterminal = 0
            else:
                nextnonterminal = 1.0 - dones[t + 1]

            delta = rewards[t] + GAMMA * (next_value[t] * nextnonterminal * (1-dones[t])) - (values[t]*(1-dones[t]))

            lastgaelam = (delta + GAMMA * LAM * lastgaelam)*(1.0-dones[t])
            advantages[t] = lastgaelam.item()

        target = advantages + values


        advantages = tf.convert_to_tensor(advantages, dtype=tf.float32)
        target = tf.convert_to_tensor(target, dtype=tf.float32)

        #print(f"advantage mean {tf.reduce_mean(advantages)}")
        #print(f"target mean {tf.reduce_mean(target)}")

        old_states = tf.stack(memory.states)
        old_states = tf.squeeze(old_states, axis=1)
        old_actions = tf.stack(memory.actions)  
        old_logprobs = tf.stack(memory.logprobs) 

        dataset = tf.data.Dataset.from_tensor_slices((old_states, old_actions, old_logprobs, advantages, target))



        for epoch in range(ppo_epochs):

            dataset_pr = dataset.shuffle(buffer_size=step).take(HORIZONT).batch(self.batch_size)

            for batch in dataset_pr:
                b_states, b_actions, b_logprobs, b_advantages, b_target = batch

                # Actor (Policy) Update
                with tf.GradientTape() as tape_actor:
                    logprobs = []
                    entropy = []
                    for state, action in zip(b_states, b_actions):
                        action_mean, std_m = self.policy(tf.convert_to_tensor([state]))
                        #std_m = tf.clip_by_value(std_m, 0.15, 1.50)
                        std_m = tf.clip_by_value(std_m, 0.15, 0.75)
                        #dist = tfp.distributions.TruncatedNormal(loc=action_mean, scale=std_m, low=-1.00, high=1.00)
                        dist = tfp.distributions.Normal(loc=action_mean, scale=std_m)
                        entropy.append(dist.entropy())
                        logprob = dist.log_prob(action)
                        logprobs.append(logprob)

                    entropy = tf.stack(entropy)
                    logprobs = tf.stack(logprobs)

                    ratio = tf.exp(logprobs - b_logprobs)
                    #print(f"ratio mean {tf.reduce_mean(ratio)}")
                    surr1 = ratio * b_advantages
                    surr2 = tf.clip_by_value(ratio, 1 - ppo_clip, 1 + ppo_clip) * b_advantages
                    loss_actor = -tf.reduce_mean(tf.minimum(surr1, surr2)) - 0.01 * tf.reduce_mean(entropy)
                    actor.append(loss_actor)

                    


                gradients_actor = tape_actor.gradient(loss_actor, self.policy.trainable_variables)


                self.optimizer.apply_gradients(zip(gradients_actor, self.policy.trainable_variables))

                # Critic (Value) Update
                with tf.GradientTape() as tape_critic:

                    state_values_list = []
                    for state in b_states:
                        state_value = self.policy_critic(tf.convert_to_tensor([state]))
                        state_values_list.append(state_value)
                    state_values = tf.stack(state_values_list)  # Tensor con tutti gli state values

                    loss_critic = 0.5 * self.MseLoss(state_values, b_target)
                    critic.append(loss_critic)


                gradients_critic = tape_critic.gradient(loss_critic, self.policy_critic.trainable_variables)


                self.optimizer2.apply_gradients(zip(gradients_critic, self.policy_critic.trainable_variables))


        self.policy_old.set_weights(self.policy.get_weights())
        self.policy.save_weights_model(weights_path_actor)
        self.policy_critic.save_weights_model(weights_path_critic)

        actor = sum(actor) / len(actor)
        critic = sum(critic) / len(critic)
        
        return actor, critic


def preprocess_state(state):
    state = np.ascontiguousarray(state, dtype=np.float32) / 255
    state = tf.convert_to_tensor(state)
    state = tf.expand_dims(state, axis=0) 
    return state



env = AirSimWrapper(input_shape=INPUT_SHAPE, ip="127.0.0.1", port=41451)

memory=Memory()

ppo = PPO_Agent(batch_size=BATCHSIZE)


j=0
e=1

init_point = random.choice(STARTING_POINTS)
state = env.reset(init_point)
state = preprocess_state(state)

ep = []
current_rewards = []
step=0

while e<tot_frame:

    action, action_logprob = ppo.select_action(state)
    j=j+1
    next_state, reward, done = env.step(action, j)
    #print(reward)
    ep.append(reward)
    next_state=preprocess_state(next_state)
    memory.next_obs.append(next_state)
    memory.states.append(state)
    memory.actions.append(tf.convert_to_tensor(action))
    memory.logprobs.append(action_logprob)
    memory.rewards.append(reward)
    memory.dones.append(done)
    state=next_state

    if done==1:
        init_point = random.choice(STARTING_POINTS)
        state=env.reset(init_point)
        state=preprocess_state(state)
        current_rewards.append(sum(ep))

        if sum(ep)==0:
            e=e-1
            memory.next_obs.pop()
            memory.states.pop()
            memory.actions.pop()
            memory.logprobs.pop()
            memory.rewards.pop()
            memory.dones.pop()
            current_rewards.pop()

        ep = []
        j=0

    if e % HORIZONT == 0:

        # Aspetta fino a quando l'ultimo episodio non è terminato
        while len(memory.dones) > 0 and memory.dones[-1] == 0:
            action, action_logprob = ppo.select_action(state)
            j = j + 1
            next_state, reward, done = env.step(action, j)
            # print(reward)
            ep.append(reward)
            next_state = preprocess_state(next_state)
            memory.next_obs.append(next_state)
            memory.states.append(state)
            memory.actions.append(tf.convert_to_tensor(action))
            memory.logprobs.append(action_logprob)
            memory.rewards.append(reward)
            memory.dones.append(done)
            state = next_state

            if done == 1:
                init_point = random.choice(STARTING_POINTS)
                state = env.reset2(init_point)
                #print("finito ultimo ep")
                state = preprocess_state(state)
                current_rewards.append(sum(ep))

                if sum(ep) == 0:
                    e = e - 1
                    memory.next_obs.pop()
                    memory.states.pop()
                    memory.actions.pop()
                    memory.logprobs.pop()
                    memory.rewards.pop()
                    memory.dones.pop()
                    current_rewards.pop()

                ep = []
                j = 0

        ep = []

        init_point = random.choice(STARTING_POINTS)
        state = env.reset2(init_point)
        state = preprocess_state(state)


        step = len(memory.dones)


        a, c = ppo.update(memory, step)

        init_point = random.choice(STARTING_POINTS)
        state = env.reset(init_point)
        state = preprocess_state(state)

        memory.forget()

        if not os.path.exists(loss_file_path):
            df = pd.DataFrame(columns=['LossA', 'LossC' ])
            df.to_excel(loss_file_path, index=False)  

        a_float = a.numpy().item() if hasattr(a, 'numpy') else float(a)
        b_float = c.numpy().item() if hasattr(c, 'numpy') else float(c)
        new_data = [a_float, b_float]

        try:
            wb = load_workbook(loss_file_path)
            ws = wb['Sheet1']
            ws.append(new_data)  
            wb.save(loss_file_path)
        except PermissionError as e:
            print(f"PermissionError: {e}")



        if not os.path.exists(reward_file_path):
            df = pd.DataFrame(columns=['Reward'])
            df.to_excel(reward_file_path, index=False)  

        try:
            wb = load_workbook(reward_file_path)
            ws = wb['Sheet1']

            for reward in current_rewards:
                new_data=[reward]
                ws.append(new_data)  


            wb.save(reward_file_path)

        except PermissionError as e:
            print(f"PermissionError: {e}")

        current_rewards = []



    e=e+1
